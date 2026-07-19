import csv
import json
import os
import time
import warnings

import openpyxl
import requests
from dotenv import load_dotenv, set_key


def request_dadata(request, address):
    # Зачем делать разные url по факту для одного API, та ещё и разный формат ввода, у yandex лучше.
    # Плюс они берут данные из OpenStreetMap
    if address:
        data = [f"{request}"]
    else:
        lat, lon = map(float, reversed(request.split(",")))
        data = {
            "lat": lat,
            "lon": lon,
        }
    if address:
        url = "https://cleaner.dadata.ru/api/v1/clean/address"
    else:
        url = "https://suggestions.dadata.ru/suggestions/api/4_1/rs/geolocate/address"
    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json",
        "Authorization": f"Token {os.getenv('DADATA_API_KEY')}",
        "X-Secret": f"{os.getenv('DADATA_SECRET')}",
    }
    start = time.perf_counter()
    response = requests.post(url, headers=headers, json=data)
    elapsed = time.perf_counter() - start
    json_response = json.loads(response.text)
    qc_geo_mapping = {
        0: "Точные координаты дома",
        1: "Ближайший дом",
        2: "Улица",
        3: "Населенный пункт",
        4: "Город",
        5: "Координаты не определены",
    }
    if address:
        coords = json_response[0]["geo_lat"] + " " + json_response[0]["geo_lon"]
        qc_geo = json_response[0]["qc_geo"]
        qc_geo = str(qc_geo) + " - " + qc_geo_mapping.get(int(qc_geo), "")
    else:
        try:
            coords = json_response["suggestions"][0]["unrestricted_value"]
            qc_geo = json_response["suggestions"][0]["data"]["qc_geo"]
            qc_geo = str(qc_geo) + " - " + qc_geo_mapping.get(int(qc_geo), "")
        except:
            coords = ""
            qc_geo = ""
            pass
        
        

    return json_response, coords, qc_geo, elapsed


def request_gigachat(request, address):
    if address:
        ask = f"Выведи координаты по адресу {request}, выдай только координаты без дополнительных сообщений, даже в скобках."
    else:
        ask = f"Выведи адрес по координатам в формате lon,lat {request} в виде Индекс, город, адрес, без дополнительных сообщений, даже в скобках."
    url = "https://gigachat.devices.sberbank.ru/api/v1/chat/completions"
    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json",
        "Authorization": f"Bearer {os.getenv('GIGACHAT_TOKEN')}",
    }
    payload = {
        "model": "GigaChat-2",
        "messages": [{"role": "user", "content": f"{ask}"}],
        "n": 1,
        "stream": False,
        "max_tokens": 512,
        "repetition_penalty": 1,
        "update_interval": 0,
    }
    # Сбер использует self-singed сертификат, что я считаю минусом
    # Необходимость постоянного запроса ключа раз в 30 минут, да можно сделать таймер основанный на глобальном получение времени и он будет "сохранятся" без проблем даже при нескольких запусках программы
    start = time.perf_counter()
    response = requests.post(url, headers=headers, json=payload, verify=False)
    elapsed = time.perf_counter() - start
    json_response = json.loads(response.text)
    return json_response, json_response["choices"][0]["message"]["content"], "", elapsed


def request_gigachat_access():
    if int(os.getenv("GIGACHAT_TIME") or 0) > int(time.time()):
        return
    url = "https://ngw.devices.sberbank.ru:9443/api/v2/oauth"
    payload = {"scope": "GIGACHAT_API_PERS"}
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json",
        "RqUID": "dc05e37e-9f51-4a94-a518-e3aea4c5aba0",
        "Authorization": f"Basic {os.getenv('GIGACHAT_API_KEY')}",
    }

    response = requests.request(
        "POST", url, headers=headers, data=payload, verify=False
    )
    json_response = json.loads(response.text)
    set_key(".env", "GIGACHAT_TOKEN", json_response["access_token"])
    set_key(".env", "GIGACHAT_TIME", str(json_response["expires_at"])[:-3])
    print("Token updated for 30 minutes.")


def request_yandex(request, address):
    if address:
        request = [addr.replace(" ", "+") for addr in request]
    start = time.perf_counter()
    response = requests.get(
        f"https://geocode-maps.yandex.ru/v1/?apikey={os.getenv('YANDEX_API_KEY')}&geocode={request}&format=json"
    )
    elapsed = time.perf_counter() - start
    json_response = json.loads(response.text)

    if address:
        geo_object = json_response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject']  # fmt: off
        pos = " ".join(geo_object["Point"]["pos"].split()[::-1])
        precision = geo_object["metaDataProperty"]["GeocoderMetaData"]["precision"]
        return json_response, pos, precision, elapsed
    else:
        geo_object = json_response["response"]["GeoObjectCollection"]["featureMember"][0]["GeoObject"]  # fmt: off
        try:
            postal = geo_object["metaDataProperty"]["GeocoderMetaData"]["Address"]["postal_code"]  # fmt: off
        except KeyError:
            postal = ""
        address = (
            postal + ", " + geo_object["metaDataProperty"]["GeocoderMetaData"]["text"]
        )
        precision = geo_object["metaDataProperty"]["GeocoderMetaData"]["precision"]
        return json_response, address, precision, elapsed


def write_to_log(request, response, type):
    with open("responce.txt", "a", encoding="UTF-8") as file:
        file.write(str(request) + " " + f"{type}" + "\n")
        json.dump(response[0], file, ensure_ascii=False, indent=2)
        file.write("\n")
    if not os.path.isfile("responce.csv"):
        with open("responce.csv", "w", newline="", encoding="UTF-8") as file:
            writer = csv.writer(file)
            writer.writerow(
                ["service", "request", "reponse", "quality (if exits)", "time elapsed"]
            )
    with open("responce.csv", "a", newline="", encoding="utf-8") as file:
        writer = csv.writer(file)
        writer.writerow([type, request, response[1], response[2], response[3]])


def read_excel(file):
    wb = openpyxl.load_workbook(file)
    ws1 = wb["Адреса"]
    ws2 = wb["Координаты"]
    addresses = [list(row) for row in ws1.iter_rows(min_row=2, values_only=True)]
    coordinates = [list(row) for row in ws2.iter_rows(min_row=2, values_only=True)]
    x_coordinates = [float(str(item[0]).split(", ")[0]) for item in coordinates]
    y_coordinates = [float(str(item[0]).split(", ")[1]) for item in coordinates]
    return addresses, (x_coordinates, y_coordinates)


if __name__ == "__main__":
    warnings.filterwarnings("ignore")
    load_dotenv()
    request_gigachat_access()
    addresses, coordinates = read_excel("Геоданные.xlsx")

    for address in addresses:
        response = request_yandex(address, True)
        write_to_log(address, response, "YANDEX")
        response = request_gigachat(address, True)
        write_to_log(address, response, "GIGACHAT")
        response = request_dadata(address, True)
        write_to_log(address, response, "DADATA")
        time.sleep(1)

    for i in range(len(coordinates[0])):
        coords = f"{coordinates[1][i]},{coordinates[0][i]}"
        response = request_yandex(coords, False)
        write_to_log(coords, response, "YANDEX")
        response = request_gigachat(coords, False)
        write_to_log(coords, response, "GIGACHAT")
        response = request_dadata(coords, False)
        write_to_log(coords, response, "DADATA")
        time.sleep(1)
